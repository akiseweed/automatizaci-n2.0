from PySide2 import QtWidgets , QtCore , QtGui
from PySide2.QtCore import Qt


from controllers.modulo_utils.credenciales_utils import evento_tabla, llenar_tabla_automatizacion, llenar_tabla_credenciales
from db.Modulo_base import BaseDatos
from view.ui_principal import Ui_MainWindow

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from threading import Thread
import time
import re, os
from datetime import datetime
import xlsxwriter
import os
from PySide2.QtWidgets import QFileDialog, QMessageBox


from openpyxl.styles import Alignment, Font, PatternFill
import openpyxl


class Principal(QtWidgets.QMainWindow):
    
    def __init__(self, nombreusuario= "", idusuario= None , parent=None):
        super(Principal, self).__init__(parent)
        self.venPri = Ui_MainWindow()
        self.venPri.setupUi(self)
        
        self.base = BaseDatos()
        
        evento_tabla(self)
        
        self.idUsuario  = idusuario
        self.nombre = nombreusuario
        
        self.venPri.nombreusuario.setText(""+ nombreusuario)
        self.llenarTablaCredenciales(self.idUsuario)
         
        self.detener_automatizacion = False
        
        # boton que inicia la automatizacion
        self.venPri.btn_iniciar.clicked.connect(lambda:self.iniciarAutomatizacion())
        self.venPri.btn_finalizar.clicked.connect(lambda:self.finalizarAutomatizacion())
        
        self.venPri.btn_cerrarsesion.clicked.connect(lambda:self.cerrarSesion())
        
        self.venPri.line_busqueda.textChanged.connect(
            lambda: self.busqueda_credenciale(self.venPri.line_busqueda.text()))
        
        self.venPri.btn_export.clicked.connect(lambda: self.exportarExcel())
        self.venPri.btn_mover.clicked.connect(lambda: self.pasarlote())
    
    
    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_Escape:
            pass
        
    def llenarTablaCredenciales(self, idUsuario):
        # Consulta para obtener todas las credenciales asociadas al usuario con el ID proporcionado
        consulta = """
            SELECT id, numero_empresa, razon_social, ruc, usuario, clave
            FROM credenciales 
            JOIN credenciales_usuario ON credenciales.id = credenciales_usuario.credenciales_id 
            WHERE credenciales_usuario.usuario_id = %s
        """
        
        # Luego ejecutas la consulta con el ID de usuario como parámetro
        resultado = self.base.getDatos_condicion(consulta, [idUsuario])
        
        if(resultado):
            llenar_tabla_credenciales(self, self.venPri.tabla_credenciales, resultado)
        
    def busqueda_credenciale(self, texto):
        
        if len(texto) != 0:
            # Consulta para buscar registros en la tabla credenciales que coincidan con el texto ingresado
            consulta = """
                SELECT * 
                FROM credenciales 
                WHERE razon_social LIKE %s 
                    OR numero_empresa LIKE %s 
                    OR ruc LIKE %s 
                    OR usuario LIKE %s
            """
            
            # Parámetros para el operador LIKE
            texto_like = f"%{texto}%"
            
            # Ejecutar la consulta con los parámetros
            resultado = self.base.getDatos_condicion(consulta, [texto_like, texto_like, texto_like, texto_like])
            
            if(resultado):
                llenar_tabla_credenciales(self, self.venPri.tabla_credenciales, resultado)
            else:
                self.llenarTablaCredenciales(self.idUsuario)
         

    def exportarExcel(self):
        consulta = """
            SELECT id, numero_empresa, razon_social, , ruc, usuario, clave
            FROM credenciales 
            JOIN credenciales_usuario ON credenciales.id = credenciales_usuario.credenciales_id 
            WHERE credenciales_usuario.usuario_id = %s
        """

        resultado = self.base.getDatos_condicion(consulta, [self.idUsuario])
        
        if resultado:
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            titulo_celda = ws.cell(row=1, column=3)
            titulo_celda.value = "CREDENCIALES - " + self.nombre
            titulo_celda.alignment = Alignment(horizontal='center')
            titulo_celda.font = Font(bold=True)
            
        
            cabecera = ["ID", "RAZON SOCIAL", "NUMERO EMPRESA", "RUC", "USUARIO", "CLAVE"]
            ws.append(cabecera)

            for cell in ws[1]:
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                        
            for fila in resultado:
                ws.append(list(fila))
            
            for columna in ws.columns:
                longitud_maxima = 0
                for celda in columna:
                    if celda.value:
                        longitud = len(str(celda.value))
                        if longitud > longitud_maxima:
                            longitud_maxima = longitud
                ajuste = (longitud_maxima + 2) * 1.2 
                ws.column_dimensions[columna[0].column_letter].width = ajuste
            
            dialog = QFileDialog(self)
            dialog.setWindowTitle("Guardar como")
            dialog.setAcceptMode(QFileDialog.AcceptSave)
            dialog.setNameFilter("Archivos de Excel (*.xlsx)")
            
            if dialog.exec_() == QFileDialog.Accepted:
                ruta_guardado = dialog.selectedFiles()[0]
                
                wb.save(ruta_guardado)
            
                QMessageBox.information(self, "Guardado", "El archivo se ha guardado correctamente.")


    def pasarlote(self):

        datos_tabla_origen = []
        for fila in range(self.venPri.tabla_credenciales.rowCount()):
            fila_datos = []
            for columna in range(self.venPri.tabla_credenciales.columnCount()):
                item = self.venPri.tabla_credenciales.item(fila, columna)
                if item is not None:
                    fila_datos.append(item.text())
                else:
                    fila_datos.append("")
            datos_tabla_origen.append(fila_datos)

        llenar_tabla_automatizacion(self, self.venPri.tabla_automatizar, datos_tabla_origen)

        
    def cerrarSesion(self):
        from controllers.modulo_login.login import Login
        
        self.close()
        self.login = Login()
        self.login.show()
    
    def iniciarAutomatizacion(self):
        
        self.automatizacion_thread = Thread(target=self.automatizarCredenciales)
        self.automatizacion_thread.start()    
        
        
    def automatizarCredenciales(self):
         
        tabla_automatizar = self.venPri.tabla_automatizar
        tipo_trabajar = self.venPri.cbo_tipo.currentText()

        credenciales_lista = []

        for fila in range(tabla_automatizar.rowCount()):
            ruc = tabla_automatizar.item(fila, 3).text()
            usuario = tabla_automatizar.item(fila, 4).text()
            contrasena = tabla_automatizar.item(fila, 5).text()

            credenciales_lista.append([ruc, usuario, contrasena])
            

        sesiones = []
        for credencial in credenciales_lista:
            sesion = {
                'ruc': credencial[0],
                'usuario': credencial[1],
                'contrasena': credencial[2]
            }
            sesiones.append(sesion)
            
        controladores = []
        for sesion in sesiones:
            
            if(tipo_trabajar == "Sunat"):
                self.iniciar_sesion_sunat(**sesion, controladores=controladores)
            elif(tipo_trabajar == 'Sunafil'):
                self.iniciar_sesion_sunafil(**sesion, controladores=controladores)
            else:
                self.iniciar_sesion_AFPnet(**sesion, controladores=controladores)
            
        self.detener_automatizacion = False
        
        while not self.detener_automatizacion:
            pass


    def iniciar_sesion_sunat(self, ruc, usuario, contrasena, controladores):

        driver = webdriver.Chrome()

        login_url = 'https://api-seguridad.sunat.gob.pe/v1/clientessol/4f3b88b3-d9d6-402a-b85d-6a0bc857746a/oauth2/loginMenuSol?originalUrl=https://e-menu.sunat.gob.pe/cl-ti-itmenu/AutenticaMenuInternet.htm&state=rO0ABXNyABFqYXZhLnV0aWwuSGFzaE1hcAUH2sHDFmDRAwACRgAKbG9hZEZhY3RvckkACXRocmVzaG9sZHhwP0AAAAAAAAx3CAAAABAAAAADdAAEZXhlY3B0AAZwYXJhbXN0AEsqJiomL2NsLXRpLWl0bWVudS9NZW51SW50ZXJuZXQuaHRtJmI2NGQyNmE4YjVhZjA5MTkyM2IyM2I2NDA3YTFjMWRiNDFlNzMzYTZ0AANleGVweA==%20%27'

        driver.get(login_url)

        ruc_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'btnPorRuc')))
        ruc_button.click()

        ruc_input = driver.find_element(By.ID, 'txtRuc')
        ruc_input.send_keys(ruc)
        usuario_input = driver.find_element(By.ID, 'txtUsuario')
        usuario_input.send_keys(usuario)
        contrasena_input = driver.find_element(By.ID, 'txtContrasena')
        contrasena_input.send_keys(contrasena)
        submit_button = driver.find_element(By.ID, 'btnAceptar')
        submit_button.click()
        
        submit_button = driver.find_element(By.ID, 'aOpcionBuzon')
        submit_button.click()

        time.sleep(3)

        controladores.append(driver)


    def iniciar_sesion_sunafil(self, ruc, usuario, contrasena, controladores):
        driver = webdriver.Chrome()
        
        login_url_sunafil = 'https://api-seguridad.sunat.gob.pe/v1/clientessol/b6474e23-8a3b-4153-b301-dafcc9646250/oauth2/login?originalUrl=https://casillaelectronica.sunafil.gob.pe/si.inbox/Login/Empresa&state=s'
        
        driver.get(login_url_sunafil)
        
        ruc_input = driver.find_element(By.ID, 'txtRuc')
        ruc_input.send_keys(ruc)

        usuario_input = driver.find_element(By.ID, 'txtUsuario')
        usuario_input.send_keys(usuario)

        contrasena_input = driver.find_element(By.ID, 'txtContrasena')
        contrasena_input.send_keys(contrasena)

        submit_button = driver.find_element(By.ID, 'btnAceptar')
        submit_button.click()

        time.sleep(2)

        controladores.append(driver)

    
    def iniciar_sesion_AFPnet(self, ruc, usuario, contrasena, controladores):

        driver = webdriver.Chrome()

        # Abrir la página de AFPNet
        driver.get("https://www.afpnet.com.pe/")

   
        ruc_input = driver.find_element(By.ID, 'NumeroDocumento')
        ruc_input.send_keys(ruc)  

        usuario_input = driver.find_element(By.ID, 'NombreUsuario')
        usuario_input.send_keys(usuario)  
        
        

        # #keypad > div:nth-child(1) > div > button:nth-child(1)
        # #keypad > div:nth-child(1) > div > button:nth-child(2)
        # #keypad > div:nth-child(1) > div > button:nth-child(3)
        # #keypad > div:nth-child(1) > div > button:nth-child(4)
        # #keypad > div:nth-child(1) > div > button:nth-child(5)
        # #keypad > div:nth-child(1) > div > button:nth-child(6)
        # #keypad > div:nth-child(1) > div > button:nth-child(7)
        # #keypad > div:nth-child(1) > div > button:nth-child(8)
        # #keypad > div:nth-child(1) > div > button:nth-child(0)

        
        usuario_contrasena = driver.find_element(By.ID, 'txtContrasenia')
        usuario_contrasena.send_keys(contrasena)  
        
        
         # Agregar el controlador a la lista
        controladores.append(driver)


    def finalizarAutomatizacion(self):
        self.detener_automatizacion = True
        llenar_tabla_automatizacion(self, self.venPri.tabla_automatizar, [])